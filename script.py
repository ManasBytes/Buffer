import math
import os
import traceback
import warnings
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


warnings.filterwarnings(
    "ignore",
    message="Workbook contains no default style, apply openpyxl's default",
    category=UserWarning,
)


MAIN_INPUT_FOLDER = Path("main")
OUTPUT_ROOT_FOLDER = Path("output")
CANDIDATES_DIR_NAME = "candidates"
SHEET_PROTECTION_PASSWORD = "QP_QuestionPaper_ExcelDocument_SheetProtection_2026_9f3c7a1e4d2b8c6f5a0d7e1c9b4f6a8d"


def normalize_lookup(value):
    return "".join(ch.lower() for ch in str(value).strip() if ch.isalnum())


def find_column(columns, candidates):
    normalized_map = {normalize_lookup(column): column for column in columns}
    normalized_candidates = [normalize_lookup(candidate) for candidate in candidates]

    for normalized_candidate in normalized_candidates:
        if normalized_candidate in normalized_map:
            return normalized_map[normalized_candidate]

    for normalized_candidate in normalized_candidates:
        for normalized_column, original_column in normalized_map.items():
            if normalized_column.startswith(normalized_candidate) or normalized_candidate.startswith(normalized_column):
                return original_column

    return None


def normalize_id(value):
    if pd.isna(value):
        return ""
    try:
        return str(int(float(value)))
    except (ValueError, TypeError):
        return str(value).strip()


def normalize_question_id_for_match(value):
    return normalize_id(value).strip()


def option_token_to_number(token):
    normalized_token = str(token).strip().lower()
    if normalized_token == "":
        return ""

    if len(normalized_token) == 1 and "a" <= normalized_token <= "z":
        return str(ord(normalized_token) - ord("a") + 1)

    try:
        return str(int(float(normalized_token)))
    except (ValueError, TypeError):
        return ""


def concatenate_options(option_values, separator=","):
    ordered_non_empty = [value for value in option_values if value != ""]
    return separator.join(ordered_non_empty)


def parse_correct_option(value):
    if pd.isna(value):
        return ""
    text = str(value).strip()
    if text == "":
        return ""

    normalized_values = []
    for token in text.split(","):
        option_number = option_token_to_number(token)
        if option_number == "":
            continue
        normalized_values.append(option_number)

    return ",".join(normalized_values)


def get_correct_option_id(concatenated_option_string, correct_option_number, separator=","):
    if correct_option_number == "" or pd.isna(correct_option_number):
        return ""

    option_ids = [item for item in str(concatenated_option_string).split(separator) if item != ""]
    correct_option_tokens = [token.strip() for token in str(correct_option_number).split(separator) if token.strip() != ""]

    mapped_correct_option_ids = []
    for token in correct_option_tokens:
        try:
            index = int(float(token)) - 1
        except (ValueError, TypeError):
            return ""
        if index < 0 or index >= len(option_ids):
            return ""
        mapped_correct_option_ids.append(option_ids[index])

    return separator.join(mapped_correct_option_ids)


def normalize_candidate_entry(value):
    if pd.isna(value):
        return ""
    return str(value).strip().lower()


def candidate_entry_to_redirect(candidate_entry, question_type):
    candidate_text = normalize_candidate_entry(candidate_entry)
    if candidate_text == "":
        return ""

    if question_type in ("MCQ", "MSQ"):
        mapped_values = []
        for token in candidate_text.split(","):
            option_number = option_token_to_number(token)
            if option_number == "":
                continue
            mapped_values.append(option_number)
        return ",".join(mapped_values)

    if question_type == "SA":
        return candidate_text

    return candidate_text


def redirect_to_selected_id(concatenated_option_string, redirect_to_number, separator=","):
    redirect_text = normalize_candidate_entry(redirect_to_number)
    if redirect_text == "":
        return "", False

    option_ids = [item for item in str(concatenated_option_string).split(separator) if item != ""]
    selected_ids = []
    for token in redirect_text.split(separator):
        token = token.strip()
        if token == "":
            continue
        try:
            option_index = int(token) - 1
        except ValueError:
            return "", True
        if option_index < 0 or option_index >= len(option_ids):
            return "", True
        selected_ids.append(option_ids[option_index])

    return separator.join(selected_ids), False


def normalize_marks(value):
    if pd.isna(value):
        return ""
    try:
        numeric_value = float(value)
        if numeric_value.is_integer():
            return int(numeric_value)
        return numeric_value
    except (ValueError, TypeError):
        return str(value).strip()


def normalize_sa_answer(value):
    if pd.isna(value):
        return ""
    return str(value).strip()


def parse_yes_no(value):
    text = "" if pd.isna(value) else str(value).strip().lower()
    if text in ("yes", "y", "true", "1"):
        return True
    if text in ("no", "n", "false", "0"):
        return False
    return None


def split_sa_answer_tokens(answer_text, separator="<sa_ans_sep>"):
    return [token.strip() for token in str(answer_text).split(separator) if token.strip() != ""]


def parse_numeric_token(value):
    try:
        return float(str(value).strip())
    except (ValueError, TypeError):
        return None


def numeric_equal(left, right, rel_tol=1e-9, abs_tol=1e-9):
    return math.isclose(left, right, rel_tol=rel_tol, abs_tol=abs_tol)


def normalize_sa_text(value, case_sensitive):
    text = "" if pd.isna(value) else str(value).strip()
    return text if case_sensitive else text.lower()


def evaluate_sa_answer(candidate_entry, correct_answer, marks, response_type, evaluation_required, answer_type, case_sensitive):
    evaluation_required_value = parse_yes_no(evaluation_required)
    if evaluation_required_value is False:
        return 0, "M"

    candidate_text = "" if pd.isna(candidate_entry) else str(candidate_entry).strip()
    if candidate_text == "":
        return 0, "U"

    correct_answer_text = "" if pd.isna(correct_answer) else str(correct_answer).strip()
    if correct_answer_text == "":
        return 0, "W"

    response_type_text = "" if pd.isna(response_type) else str(response_type).strip().lower()
    answer_type_text = "" if pd.isna(answer_type) else str(answer_type).strip().lower()
    case_sensitive_value = parse_yes_no(case_sensitive) is True

    if response_type_text == "numeric":
        candidate_number = parse_numeric_token(candidate_text)
        if candidate_number is None:
            return 0, "W"

        if answer_type_text == "range":
            range_tokens = split_sa_answer_tokens(correct_answer_text)
            if len(range_tokens) != 2:
                return 0, "W"
            range_start = parse_numeric_token(range_tokens[0])
            range_end = parse_numeric_token(range_tokens[1])
            if range_start is None or range_end is None:
                return 0, "W"
            lower_bound = min(range_start, range_end)
            upper_bound = max(range_start, range_end)
            if lower_bound <= candidate_number <= upper_bound:
                return marks, "C"
            return 0, "W"

        if answer_type_text == "equal":
            expected_number = parse_numeric_token(correct_answer_text)
            if expected_number is None:
                return 0, "W"
            if numeric_equal(candidate_number, expected_number):
                return marks, "C"
            return 0, "W"

        if answer_type_text == "set":
            expected_tokens = split_sa_answer_tokens(correct_answer_text)
            expected_numbers = [parse_numeric_token(token) for token in expected_tokens]
            expected_numbers = [value for value in expected_numbers if value is not None]
            if len(expected_numbers) == 0:
                return 0, "W"
            if any(numeric_equal(candidate_number, expected_number) for expected_number in expected_numbers):
                return marks, "C"
            return 0, "W"

        return 0, "W"

    if response_type_text == "alphanumeric":
        candidate_normalized = normalize_sa_text(candidate_text, case_sensitive_value)

        if answer_type_text == "equal":
            expected_normalized = normalize_sa_text(correct_answer_text, case_sensitive_value)
            if candidate_normalized == expected_normalized:
                return marks, "C"
            return 0, "W"

        if answer_type_text == "set":
            expected_tokens = split_sa_answer_tokens(correct_answer_text)
            expected_normalized_tokens = [normalize_sa_text(token, case_sensitive_value) for token in expected_tokens]
            if candidate_normalized in expected_normalized_tokens:
                return marks, "C"
            return 0, "W"

        return 0, "W"

    return 0, "W"


def split_non_empty_csv(value, separator=","):
    return [token.strip() for token in str(value).split(separator) if token.strip() != ""]


def evaluate_correct_mark_and_result(question_type, selected_id, correct_option_id, marks, invalid_redirect=False, sa_response_type="", sa_evaluation_required="", sa_answer_type="", sa_case_sensitive=""):
    if question_type == "SA":
        return evaluate_sa_answer(
            candidate_entry=selected_id,
            correct_answer=correct_option_id,
            marks=marks,
            response_type=sa_response_type,
            evaluation_required=sa_evaluation_required,
            answer_type=sa_answer_type,
            case_sensitive=sa_case_sensitive,
        )

    if question_type not in ("MCQ", "MSQ"):
        return "", ""

    if invalid_redirect:
        return 0, "W"

    selected_tokens = split_non_empty_csv(selected_id)
    correct_tokens = split_non_empty_csv(correct_option_id)

    if len(selected_tokens) == 0:
        return 0, "U"

    if len(correct_tokens) == 0:
        return 0, "W"

    if question_type == "MCQ":
        if len(selected_tokens) == 1 and len(correct_tokens) == 1 and selected_tokens[0] == correct_tokens[0]:
            return marks, "C"
        return 0, "W"

    selected_set = set(selected_tokens)
    correct_set = set(correct_tokens)

    if selected_set == correct_set and len(selected_tokens) == len(correct_tokens):
        return marks, "C"

    if selected_set.issubset(correct_set):
        return 0, "PC"

    return 0, "W"


def resolve_sheet_name(sheet_names, candidates):
    normalized_sheet_map = {normalize_lookup(sheet_name): sheet_name for sheet_name in sheet_names}
    normalized_candidates = [normalize_lookup(candidate) for candidate in candidates]

    for normalized_candidate in normalized_candidates:
        if normalized_candidate in normalized_sheet_map:
            return normalized_sheet_map[normalized_candidate]

    for normalized_candidate in normalized_candidates:
        for normalized_sheet_name, original_sheet_name in normalized_sheet_map.items():
            if normalized_sheet_name.startswith(normalized_candidate) or normalized_candidate.startswith(normalized_sheet_name):
                return original_sheet_name

    return None


def build_candidate_entry_map(answer_sheet_file_path):
    candidate_entry_map = {}
    answer_question_id_sequence = []

    if str(answer_sheet_file_path).strip() == "" or not os.path.exists(answer_sheet_file_path):
        return candidate_entry_map, answer_question_id_sequence

    answer_sheets = pd.read_excel(answer_sheet_file_path, sheet_name=None)
    question_paper_sheet_name = resolve_sheet_name(answer_sheets.keys(), ["Question Paper Details", "QuestionPaperDetails"])
    if question_paper_sheet_name is None:
        return candidate_entry_map, answer_question_id_sequence

    answer_df = answer_sheets[question_paper_sheet_name]
    question_id_column = find_column(answer_df.columns, ["Question id", "Question ID", "QuestionId"])
    answer_column = find_column(answer_df.columns, ["Enter your answer", "EnterYourAnswer", "Candidate Entry", "Answer"])

    if question_id_column is None or answer_column is None:
        return candidate_entry_map, answer_question_id_sequence

    answer_question_ids = answer_df[question_id_column].apply(normalize_question_id_for_match)
    answer_values = answer_df[answer_column].apply(lambda value: "" if pd.isna(value) else str(value))

    for question_id, answer_value in zip(answer_question_ids, answer_values):
        if question_id != "":
            answer_question_id_sequence.append(question_id)
            candidate_entry_map[question_id] = answer_value

    return candidate_entry_map, answer_question_id_sequence


def normalize_basic_details_value(value):
    if pd.isna(value):
        return ""
    return str(value).strip()


def extract_basic_detail_value(basic_df, label, all_labels):
    target_key = normalize_lookup(label)
    label_keys = {normalize_lookup(item) for item in all_labels}

    row_count, col_count = basic_df.shape
    for row_index in range(row_count):
        for col_index in range(col_count):
            cell_key = normalize_lookup(basic_df.iat[row_index, col_index])
            if cell_key != target_key:
                continue

            for next_col_index in range(col_index + 1, col_count):
                candidate = normalize_basic_details_value(basic_df.iat[row_index, next_col_index])
                if candidate != "":
                    return candidate

            next_row_index = row_index + 1
            if next_row_index < row_count:
                below_same_col = normalize_basic_details_value(basic_df.iat[next_row_index, col_index])
                if below_same_col != "" and normalize_lookup(below_same_col) not in label_keys:
                    return below_same_col

                for any_col_index in range(col_count):
                    candidate = normalize_basic_details_value(basic_df.iat[next_row_index, any_col_index])
                    if candidate != "" and normalize_lookup(candidate) not in label_keys:
                        return candidate

            return ""

    return ""


def build_basic_details_rows(answer_sheet_file_path):
    if str(answer_sheet_file_path).strip() == "" or not os.path.exists(answer_sheet_file_path):
        return []

    answer_sheets = pd.read_excel(answer_sheet_file_path, sheet_name=None, header=None)
    basic_details_sheet_name = resolve_sheet_name(answer_sheets.keys(), ["Basic Details", "BasicDetails"])
    if basic_details_sheet_name is None:
        return []

    basic_df = answer_sheets[basic_details_sheet_name]
    labels = ["Name", "DOB", "Roll No"]
    return [(label, extract_basic_detail_value(basic_df, label, labels)) for label in labels]


def format_total_value(value):
    numeric_value = float(value)
    if numeric_value.is_integer():
        return int(numeric_value)
    return numeric_value


def sum_numeric_values(values):
    numeric_values = [parse_numeric_token(value) for value in values]
    numeric_values = [value for value in numeric_values if value is not None]
    return sum(numeric_values)


def build_marks_rows(result_values, marks_values, correct_mark_values):
    total_correct = sum(1 for value in result_values if str(value).strip() == "C")
    total_partial_correct = sum(1 for value in result_values if str(value).strip() == "PC")
    total_manual_corrections = sum(1 for value in result_values if str(value).strip() == "M")
    total_wrong = sum(1 for value in result_values if str(value).strip() == "W")

    total_marks = format_total_value(sum_numeric_values(marks_values))
    total_correct_marks = format_total_value(sum_numeric_values(correct_mark_values))

    return [
        ("Marks", ""),
        ("Total Correct", total_correct),
        ("Total Partial Corrects", total_partial_correct),
        ("Total Manual Corrections", total_manual_corrections),
        ("Total Wrongly Answered", total_wrong),
        ("Total Marks", total_marks),
        ("Total Correct Marks", total_correct_marks),
    ]


def build_side_panel_columns(total_rows, side_panel_rows):
    basic_detail_labels = [""] * total_rows
    basic_detail_values = [""] * total_rows

    for row_index, (label, value) in enumerate(side_panel_rows):
        if row_index >= total_rows:
            break
        basic_detail_labels[row_index] = label
        basic_detail_values[row_index] = value

    return basic_detail_labels, basic_detail_values


def sanitize_sheet_name(raw_name, used_names):
    fallback = "Student"
    value = str(raw_name).strip() if raw_name is not None else ""
    if value == "":
        value = fallback

    invalid_chars = set('[]:*?/\\')
    cleaned = "".join(ch for ch in value if ch not in invalid_chars).strip()
    if cleaned == "":
        cleaned = fallback

    cleaned = cleaned[:31]
    unique_name = cleaned
    suffix = 1
    while unique_name in used_names:
        suffix_text = f"_{suffix}"
        unique_name = cleaned[: max(0, 31 - len(suffix_text))] + suffix_text
        suffix += 1

    used_names.add(unique_name)
    return unique_name


def get_config_and_english_frames(config_file_path):
    all_sheets = pd.read_excel(config_file_path, sheet_name=None)

    configuration_sheet_name = resolve_sheet_name(all_sheets.keys(), ["Configuration Details", "ConfigurationDetail"])
    english_sheet_name = resolve_sheet_name(all_sheets.keys(), ["English"])

    if configuration_sheet_name is None:
        raise KeyError("Configuration Details sheet not found")
    if english_sheet_name is None:
        raise KeyError("English sheet not found")

    config_df = all_sheets[configuration_sheet_name]
    english_df = all_sheets[english_sheet_name]
    return config_df, english_df


def build_candidate_display_df(config_df, english_df, answersheet_path):
    config_question_id_col = find_column(config_df.columns, ["Question id", "Question ID", "QuestionId"])
    config_question_type_col = find_column(config_df.columns, ["Question Type", "QuestionType"])
    config_marks_col = find_column(config_df.columns, ["Marks", "Mark"])

    if config_question_id_col is None or config_question_type_col is None or config_marks_col is None:
        raise KeyError("Missing required columns in Configuration Details")

    filtered_df = config_df[config_df[config_question_type_col].isin(["MCQ", "MSQ", "SA"])].copy()

    all_columns = list(config_df.columns)
    no_of_options_col = find_column(config_df.columns, ["No Of Options", "NoOfOptions"])
    if no_of_options_col is None:
        raise KeyError("Missing required column in Configuration Details: No Of Options")
    no_of_options_index = all_columns.index(no_of_options_col)
    option_source_columns = all_columns[no_of_options_index + 1:no_of_options_index + 5]
    if len(option_source_columns) < 4:
        raise KeyError("Option ID columns not found after 'No Of Options'")

    option_1 = filtered_df[option_source_columns[0]].apply(normalize_id)
    option_2 = filtered_df[option_source_columns[1]].apply(normalize_id)
    option_3 = filtered_df[option_source_columns[2]].apply(normalize_id)
    option_4 = filtered_df[option_source_columns[3]].apply(normalize_id)

    separator_value = ","
    concatenated_options = [
        concatenate_options([o1, o2, o3, o4], separator_value)
        for o1, o2, o3, o4 in zip(option_1, option_2, option_3, option_4)
    ]

    english_question_id_col = find_column(english_df.columns, ["Question ID", "Question id", "QuestionId"])
    english_correct_option_col = find_column(english_df.columns, ["Correct Option", "CorrectOption"])
    english_answer_col = find_column(
        english_df.columns,
        [
            "Answer(For SA)/Skeletal Code(For Programming Test)/Static text (For Typing Test)",
            "Answer(For SA)/\nSkeletal Code(For Programming Test)/\nStatic text (For Typing Test)",
            "Answer(For SA)",
        ],
    )

    if english_question_id_col is None or english_correct_option_col is None:
        raise KeyError("Missing required columns in English sheet")

    english_question_ids = english_df[english_question_id_col].apply(normalize_question_id_for_match)
    english_correct_options = english_df[english_correct_option_col].apply(parse_correct_option)
    correct_option_map = dict(zip(english_question_ids, english_correct_options))

    english_sa_answers = (
        english_df[english_answer_col].apply(normalize_sa_answer)
        if english_answer_col is not None
        else pd.Series([""] * len(english_df))
    )
    english_sa_answer_map = dict(zip(english_question_ids, english_sa_answers))

    sa_response_type_column = find_column(filtered_df.columns, ["Response Type (For SA type of Questions)", "Response Type"])
    sa_evaluation_required_column = find_column(filtered_df.columns, ["Is Evaluation Required (For SA type of Questions)", "Is Evaluation Required"])
    sa_answer_type_column = find_column(filtered_df.columns, ["Answer type (For SA type of Questions)", "Answer Type"])
    sa_case_sensitive_column = find_column(filtered_df.columns, ["Answers case sensitive? (For SA type of Questions)", "Answers case sensitive"])

    empty_sa_series = pd.Series([""] * len(filtered_df), index=filtered_df.index)
    sa_response_type_values = filtered_df[sa_response_type_column].fillna("") if sa_response_type_column else empty_sa_series
    sa_evaluation_required_values = filtered_df[sa_evaluation_required_column].fillna("") if sa_evaluation_required_column else empty_sa_series
    sa_answer_type_values = filtered_df[sa_answer_type_column].fillna("") if sa_answer_type_column else empty_sa_series
    sa_case_sensitive_values = filtered_df[sa_case_sensitive_column].fillna("") if sa_case_sensitive_column else empty_sa_series

    config_question_ids = filtered_df[config_question_id_col].apply(normalize_question_id_for_match)
    correct_option_values = config_question_ids.map(correct_option_map).fillna("")

    sa_question_mask = filtered_df[config_question_type_col] == "SA"
    sa_correct_answers = config_question_ids.map(english_sa_answer_map).fillna("")
    correct_option_values = correct_option_values.where(~sa_question_mask, sa_correct_answers)

    correct_option_ids = [
        get_correct_option_id(option_list, correct_option, separator_value)
        for option_list, correct_option in zip(concatenated_options, correct_option_values)
    ]

    candidate_entry_map, answer_question_id_sequence = build_candidate_entry_map(answersheet_path)
    candidate_entry_values = config_question_ids.map(candidate_entry_map).fillna("").values

    config_question_id_set = set(config_question_ids.values)
    unmatched_answer_question_ids = [
        question_id for question_id in answer_question_id_sequence if question_id not in config_question_id_set
    ]
    unmatched_unique_question_ids = sorted(set(unmatched_answer_question_ids))

    redirect_to_number_values = [
        candidate_entry_to_redirect(candidate_entry, question_type)
        for candidate_entry, question_type in zip(candidate_entry_values, filtered_df[config_question_type_col].values)
    ]
    selected_id_and_invalid_flags = [
        redirect_to_selected_id(option_list, redirect_to_number, separator_value)
        for option_list, redirect_to_number in zip(concatenated_options, redirect_to_number_values)
    ]
    selected_id_values = [item[0] for item in selected_id_and_invalid_flags]
    invalid_redirect_values = [item[1] for item in selected_id_and_invalid_flags]

    option_not_applicable = "Not Applicable"
    option_1_values = option_1.mask(sa_question_mask, option_not_applicable).values
    option_2_values = option_2.mask(sa_question_mask, option_not_applicable).values
    option_3_values = option_3.mask(sa_question_mask, option_not_applicable).values
    option_4_values = option_4.mask(sa_question_mask, option_not_applicable).values
    concatenated_options_values = [
        option_not_applicable if is_sa else option_value
        for is_sa, option_value in zip(sa_question_mask, concatenated_options)
    ]
    correct_option_id_values = [
        sa_correct_answer if is_sa else correct_option_id
        for is_sa, sa_correct_answer, correct_option_id in zip(sa_question_mask, sa_correct_answers, correct_option_ids)
    ]
    redirect_to_number_values = [
        option_not_applicable if is_sa else redirect_to_number
        for is_sa, redirect_to_number in zip(sa_question_mask, redirect_to_number_values)
    ]
    selected_id_values = [
        candidate_entry if is_sa else selected_id
        for is_sa, candidate_entry, selected_id in zip(sa_question_mask, candidate_entry_values, selected_id_values)
    ]
    invalid_redirect_values = [
        False if is_sa else invalid_redirect
        for is_sa, invalid_redirect in zip(sa_question_mask, invalid_redirect_values)
    ]

    marks_values = filtered_df[config_marks_col].apply(normalize_marks).values

    correct_mark_and_result = [
        evaluate_correct_mark_and_result(
            question_type,
            selected_id,
            correct_option_id,
            marks,
            invalid_redirect,
            sa_response_type,
            sa_evaluation_required,
            sa_answer_type,
            sa_case_sensitive,
        )
        for (
            question_type,
            selected_id,
            correct_option_id,
            marks,
            invalid_redirect,
            sa_response_type,
            sa_evaluation_required,
            sa_answer_type,
            sa_case_sensitive,
        ) in zip(
            filtered_df[config_question_type_col].values,
            selected_id_values,
            correct_option_id_values,
            marks_values,
            invalid_redirect_values,
            sa_response_type_values.values,
            sa_evaluation_required_values.values,
            sa_answer_type_values.values,
            sa_case_sensitive_values.values,
        )
    ]
    correct_mark_values = [item[0] for item in correct_mark_and_result]
    result_values = [item[1] for item in correct_mark_and_result]

    basic_details_rows = build_basic_details_rows(answersheet_path)
    marks_rows = build_marks_rows(result_values, marks_values, correct_mark_values)
    side_panel_rows = basic_details_rows + [("", "")] + marks_rows if len(basic_details_rows) > 0 else marks_rows

    basic_details_label_values, basic_details_value_values = build_side_panel_columns(len(filtered_df), side_panel_rows)

    output_df = pd.DataFrame({
        "S.No": range(1, len(filtered_df) + 1),
        "Question id": config_question_ids.values,
        "Question Type": filtered_df[config_question_type_col].values,
        "OPTION 1": option_1_values,
        "OPTION 2": option_2_values,
        "OPTION 3": option_3_values,
        "OPTION 4": option_4_values,
        "Separator": separator_value,
        "Concatenated Options": concatenated_options_values,
        "Correct Option": correct_option_values.values,
        "Correct Option Id": correct_option_id_values,
        "Candidate Entry": candidate_entry_values,
        "Redirect To Number": redirect_to_number_values,
        "Selected Id": selected_id_values,
        "Marks": marks_values,
        "Correct Mark": correct_mark_values,
        "Result": result_values,
        "__gap1": "",
        "__gap2": "",
        "__gap3": "",
        "Basic Details": basic_details_label_values,
        "Basic Details Value": basic_details_value_values,
    })

    display_df = output_df.rename(columns={"__gap1": "", "__gap2": " ", "__gap3": "  ", "Basic Details Value": ""})

    student_name = ""
    for label, value in basic_details_rows:
        if normalize_lookup(label) == normalize_lookup("Name"):
            student_name = str(value).strip()
            break

    mismatch_report = {
        "unmatched_count": len(unmatched_answer_question_ids),
        "unmatched_unique_ids": unmatched_unique_question_ids,
    }

    return display_df, side_panel_rows, student_name, mismatch_report


def write_candidate_sheet(workbook, sheet_name, display_df, side_panel_rows):
    worksheet = workbook.create_sheet(sheet_name)

    for row_data in dataframe_to_rows(display_df, index=False, header=True):
        worksheet.append(row_data)

    for row in worksheet.iter_rows():
        for cell in row:
            cell.protection = Protection(locked=True, hidden=False)

    if len(side_panel_rows) > 0:
        basic_details_col_idx = display_df.columns.get_loc("Basic Details") + 1
        basic_details_value_col_idx = basic_details_col_idx + 1

        worksheet.column_dimensions[get_column_letter(basic_details_col_idx)].width = 28
        worksheet.column_dimensions[get_column_letter(basic_details_value_col_idx)].width = 32

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        border_start_row = 1
        border_end_row = 1 + len(side_panel_rows)

        for row_idx in range(border_start_row, border_end_row + 1):
            for col_idx in (basic_details_col_idx, basic_details_value_col_idx):
                worksheet.cell(row=row_idx, column=col_idx).border = thin_border

    worksheet.protection.sheet = True
    worksheet.protection.set_password(SHEET_PROTECTION_PASSWORD)
    worksheet.protection.enable()


def find_subject_config_file(subject_dir):
    candidates_dir = subject_dir / CANDIDATES_DIR_NAME
    excel_files = [
        path for path in subject_dir.iterdir()
        if path.is_file() and path.suffix.lower() in {".xlsx", ".xls"} and not path.name.startswith("~$")
    ]

    if len(excel_files) == 0:
        raise FileNotFoundError(f"No config workbook found in {subject_dir}")

    excel_files.sort(key=lambda path: path.name.lower())
    return excel_files[0], candidates_dir


def iter_candidate_files(candidates_dir):
    if not candidates_dir.exists() or not candidates_dir.is_dir():
        return []

    files = [
        path for path in candidates_dir.iterdir()
        if path.is_file() and path.suffix.lower() in {".xlsx", ".xls"} and not path.name.startswith("~$")
    ]
    files.sort(key=lambda path: path.name.lower())
    return files


def process_subject(subject_dir, output_subject_dir):
    config_file_path, candidates_dir = find_subject_config_file(subject_dir)
    candidate_files = iter_candidate_files(candidates_dir)

    if len(candidate_files) == 0:
        raise FileNotFoundError(f"No candidate workbook found in {candidates_dir}")

    config_df, english_df = get_config_and_english_frames(config_file_path)

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    used_sheet_names = set()
    subject_unmatched_total = 0
    subject_mismatch_messages = []

    for candidate_file in candidate_files:
        display_df, side_panel_rows, student_name, mismatch_report = build_candidate_display_df(config_df, english_df, str(candidate_file))
        fallback_name = candidate_file.stem
        sheet_name = sanitize_sheet_name(student_name if student_name != "" else fallback_name, used_sheet_names)
        write_candidate_sheet(workbook, sheet_name, display_df, side_panel_rows)

        unmatched_count = mismatch_report["unmatched_count"]
        if unmatched_count > 0:
            subject_unmatched_total += unmatched_count
            unique_ids_text = ", ".join(mismatch_report["unmatched_unique_ids"])
            subject_mismatch_messages.append(
                f"Candidate file: {candidate_file.name} | Sheet: {sheet_name} | Unmatched rows: {unmatched_count} | Unmatched Question IDs: [{unique_ids_text}]"
            )

    output_subject_dir.mkdir(parents=True, exist_ok=True)
    output_workbook_path = output_subject_dir / config_file_path.name
    workbook.save(output_workbook_path)
    return output_workbook_path, len(candidate_files), subject_unmatched_total, subject_mismatch_messages


def write_error_log(output_root_folder, total_unmatched_count, log_lines):
    output_root_folder.mkdir(parents=True, exist_ok=True)
    error_log_path = output_root_folder / "error_log.txt"

    with open(error_log_path, "w", encoding="utf-8") as log_file:
        log_file.write("Question ID mismatch report\n")
        log_file.write(f"Total unmatched question-id rows: {total_unmatched_count}\n")
        log_file.write("\n")
        if len(log_lines) == 0:
            log_file.write("No mismatches found.\n")
        else:
            for line in log_lines:
                log_file.write(line + "\n")

    return error_log_path


def main():
    if not MAIN_INPUT_FOLDER.exists() or not MAIN_INPUT_FOLDER.is_dir():
        raise FileNotFoundError(f"Main input folder not found: {MAIN_INPUT_FOLDER.resolve()}")

    OUTPUT_ROOT_FOLDER.mkdir(parents=True, exist_ok=True)

    subject_dirs = [path for path in MAIN_INPUT_FOLDER.iterdir() if path.is_dir()]
    subject_dirs.sort(key=lambda path: path.name.lower())

    success_count = 0
    errors = []
    total_unmatched_count = 0
    mismatch_log_lines = []

    max_workers = min(8, max(1, len(subject_dirs)))
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_map = {
            executor.submit(process_subject, subject_dir, OUTPUT_ROOT_FOLDER / subject_dir.name): subject_dir
            for subject_dir in subject_dirs
        }

        for future in as_completed(future_map):
            subject_dir = future_map[future]
            try:
                output_workbook_path, candidate_count, subject_unmatched_total, subject_mismatch_messages = future.result()
                print(f"Processed: {subject_dir.name} -> {output_workbook_path} ({candidate_count} candidates)")
                if subject_unmatched_total > 0:
                    print(f"Unmatched question-id rows in {subject_dir.name}: {subject_unmatched_total}")

                total_unmatched_count += subject_unmatched_total
                mismatch_log_lines.extend([f"Subject: {subject_dir.name}"] + subject_mismatch_messages + [""])
                success_count += 1
            except Exception as exc:
                errors.append({
                    "subject": subject_dir.name,
                    "error": str(exc),
                    "traceback": traceback.format_exc(),
                })

    print("\nBatch complete")
    print(f"Subjects succeeded: {success_count}")
    print(f"Subjects failed: {len(errors)}")
    print(f"Total unmatched question-id rows: {total_unmatched_count}")

    error_log_path = write_error_log(OUTPUT_ROOT_FOLDER, total_unmatched_count, mismatch_log_lines)
    print(f"Mismatch report written: {error_log_path}")

    if errors:
        print("\nFailures:")
        for item in errors:
            print(f"Subject: {item['subject']}")
            print(f"Error: {item['error']}")
            print(item["traceback"])


if __name__ == "__main__":
    main()
