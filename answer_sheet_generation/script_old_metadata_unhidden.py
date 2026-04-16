from pathlib import Path

import os
import re
import traceback
import warnings
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Protection
from openpyxl.worksheet.protection import SheetProtection











warnings.filterwarnings(
	"ignore",
	message="Workbook contains no default style, apply openpyxl's default",
	category=UserWarning,
	module="openpyxl.styles.stylesheet",
)


MAIN_FOLDER = Path("main")
OUTPUT_FOLDER = Path("output")
ERROR_LOG_FILE = OUTPUT_FOLDER / "error_log.txt"


def normalize_name(value):
	return "".join(ch.lower() for ch in str(value).strip() if ch.isalnum())


def find_column(columns, candidates):
	normalized_map = {normalize_name(col): col for col in columns}
	for candidate in candidates:
		key = normalize_name(candidate)
		if key in normalized_map:
			return normalized_map[key]
	for candidate in candidates:
		key = normalize_name(candidate)
		for normalized_key, original_column in normalized_map.items():
			if normalized_key.startswith(key) or key.startswith(normalized_key):
				return original_column
	return None


def get_engine(path):
	return "xlrd" if path.suffix.lower() == ".xls" else "openpyxl"


def find_source_sheet(path):
	xl = pd.ExcelFile(path, engine=get_engine(path))
	candidate_names = [
		"Configuration Details",
		"Configuration Detail",
		"Question Paper Detail",
		"Question Paper Details",
	]

	for candidate in candidate_names:
		for sheet_name in xl.sheet_names:
			if normalize_name(sheet_name) == normalize_name(candidate):
				return sheet_name

	for sheet_name in xl.sheet_names:
		name = sheet_name.lower()
		if "configuration" in name or "question paper detail" in name:
			return sheet_name

	raise KeyError(f"No source sheet found in {path.name}. Available sheets: {xl.sheet_names}")


def find_template_details_sheet(path):
	xl = pd.ExcelFile(path, engine=get_engine(path))
	for sheet_name in xl.sheet_names:
		if normalize_name(sheet_name) == normalize_name("Template Details"):
			return sheet_name
	for sheet_name in xl.sheet_names:
		if "template" in sheet_name.lower() and "detail" in sheet_name.lower():
			return sheet_name
	raise KeyError(f"No Template Details sheet found in {path.name}. Available sheets: {xl.sheet_names}")


def clean_text_value(value):
	if pd.isna(value):
		return ""
	text = str(value).strip()
	if text.endswith(".0") and text[:-2].isdigit():
		return text[:-2]
	return text


def safe_filename_stem(value):
	text = re.sub(r"[\\/*?:\"<>|]", " ", str(value))
	text = re.sub(r"\s+", " ", text).strip()
	return text if text else "output"


def parse_subject_for_filename(subject_text):
	text = clean_text_value(subject_text)
	if ":" in text:
		_, right = text.split(":", 1)
		return safe_filename_stem(right.strip())
	return safe_filename_stem(text)


def extract_subject_name(subject_text):
	text = clean_text_value(subject_text)
	if ":" in text:
		_, right = text.split(":", 1)
		return right.strip()
	return text


def extract_subject_metadata(input_path):
	template_sheet = find_template_details_sheet(input_path)
	template_df = pd.read_excel(input_path, sheet_name=template_sheet, dtype=str, engine=get_engine(input_path))
	subject_col = find_column(template_df.columns, ["Subject"])
	if subject_col is None:
		raise KeyError("Missing required column(s) in Template Details: Subject")

	non_empty_values = [
		clean_text_value(value)
		for value in template_df[subject_col]
		if clean_text_value(value) != ""
	]
	if not non_empty_values:
		raise ValueError("Template Details sheet has no non-empty Subject value")

	subject_display_text = extract_subject_name(non_empty_values[0])
	subject_filename_stem = parse_subject_for_filename(subject_display_text)
	return subject_display_text, subject_filename_stem


def build_question_paper_dataframe(source_df):
	question_id_col = find_column(source_df.columns, ["Question id", "Question ID", "QuestionId"])
	question_type_col = find_column(source_df.columns, ["Question Type", "QuestionType"])
	marks_col = find_column(source_df.columns, ["Marks", "Mark"])

	missing = []
	if question_id_col is None:
		missing.append("Question id")
	if question_type_col is None:
		missing.append("Question Type")
	if marks_col is None:
		missing.append("Marks")
	if missing:
		raise KeyError(f"Missing required column(s): {', '.join(missing)}")

	question_type_text = source_df[question_type_col].fillna("").astype(str)
	filtered_df = source_df[~question_type_text.str.contains("comprehension", case=False, regex=False)].copy()

	output_df = pd.DataFrame({
		"Question id": filtered_df[question_id_col].apply(clean_text_value),
		"S.No": range(1, len(filtered_df) + 1),
		"Question Type": filtered_df[question_type_col].apply(clean_text_value),
		"Marks": filtered_df[marks_col].apply(clean_text_value),
		"Enter your answer": "",
	})

	return output_df


def build_basic_details_sheet(subject_display_text):
	return pd.DataFrame(
		[
			["Name", ""],
			["DOB", ""],
			["Roll No", ""],
			["Email", ""],
			["Subject", subject_display_text],
			["Exam Date", ""],
			["Session", ""],
		],
		columns=["Field", "Value"],
	)


def write_workbook(output_path, question_df, basic_details_df):
	workbook = Workbook()
	basic_sheet = workbook.active
	basic_sheet.title = "Basic Details"

	for row_index, row_values in enumerate(basic_details_df.itertuples(index=False), start=1):
		field_cell = basic_sheet.cell(row=row_index, column=1, value=row_values[0])
		value_cell = basic_sheet.cell(row=row_index, column=2, value=row_values[1])
		field_cell.protection = Protection(locked=True)
		value_cell.protection = Protection(locked=row_values[0] == "Subject")
		if row_values[0] == "Subject":
			value_cell.number_format = "@"

	question_sheet = workbook.create_sheet("Question Paper Details")
	for column_index, column_name in enumerate(question_df.columns, start=1):
		cell = question_sheet.cell(row=1, column=column_index, value=column_name)
		cell.protection = Protection(locked=True)

	for row_index, (_, row) in enumerate(question_df.iterrows(), start=2):
		for column_index, column_name in enumerate(question_df.columns, start=1):
			cell = question_sheet.cell(row=row_index, column=column_index, value=row[column_name])
			if column_name == "Enter your answer":
				cell.protection = Protection(locked=False)
			else:
				cell.protection = Protection(locked=True)
			if column_name == "Question id":
				cell.number_format = "@"

	question_sheet.column_dimensions["A"].width = 20
	question_sheet.column_dimensions["B"].width = 8
	question_sheet.column_dimensions["C"].width = 16
	question_sheet.column_dimensions["D"].width = 12
	question_sheet.column_dimensions["E"].width = 25
	basic_sheet.column_dimensions["A"].width = 16
	basic_sheet.column_dimensions["B"].width = 24

	basic_sheet.protection = SheetProtection(sheet=True, password="GGEaAzOZO^axfxHdnatitb6k!CtYDKMM%O5ASf1K3CiE3KiJph7WfJD%O@h59nig2zLfmT2OZbLdV810*m5*mS0PY6a1n71iG@ha")
	question_sheet.protection = SheetProtection(sheet=True, password="GGEaAzOZO^axfxHdnatitb6k!CtYDKMM%O5ASf1K3CiE3KiJph7WfJD%O@h59nig2zLfmT2OZbLdV810*m5*mS0PY6a1n71iG@ha")

	output_path.parent.mkdir(parents=True, exist_ok=True)
	workbook.save(output_path)


def process_workbook(input_path, output_path):
	source_sheet_name = find_source_sheet(input_path)
	source_df = pd.read_excel(input_path, sheet_name=source_sheet_name, dtype=str, engine=get_engine(input_path))
	subject_display_text, subject_filename_stem = extract_subject_metadata(input_path)
	resolved_output_path = output_path.with_name(f"{subject_filename_stem}.xlsx")
	question_df = build_question_paper_dataframe(source_df)
	basic_details_df = build_basic_details_sheet(subject_display_text)
	write_workbook(resolved_output_path, question_df, basic_details_df)
	return len(question_df), resolved_output_path


def main():
	if not MAIN_FOLDER.exists():
		raise FileNotFoundError(f"Main folder not found: {MAIN_FOLDER.resolve()}")

	OUTPUT_FOLDER.mkdir(parents=True, exist_ok=True)
	errors = []
	success_count = 0

	print("Process Started")

	for root, _, files in os.walk(MAIN_FOLDER):
		root_path = Path(root)
		relative_root = root_path.relative_to(MAIN_FOLDER)

		for filename in files:
			if filename.startswith("~$"):
				continue
			if Path(filename).suffix.lower() not in {".xlsx", ".xls"}:
				continue

			input_path = root_path / filename
			output_path = OUTPUT_FOLDER / relative_root / filename

			try:
				_, _ = process_workbook(input_path, output_path)
				success_count += 1
			except Exception as exc:
				errors.append({
					"file": str(input_path),
					"error": str(exc),
					"traceback": traceback.format_exc(),
				})

	if errors:
		with open(ERROR_LOG_FILE, "w", encoding="utf-8") as handle:
			handle.write(f"Error Log - {len(errors)} failure(s)\n{'=' * 60}\n")
			for index, error in enumerate(errors, start=1):
				handle.write(f"\nFailure #{index}\n")
				handle.write(f"File: {error['file']}\n")
				handle.write(f"Error: {error['error']}\n")
				handle.write("Traceback:\n")
				handle.write(error["traceback"] + "\n")

	print("\nBatch complete")
	print(f"Success: {success_count}")
	print(f"Failed: {len(errors)}")
	if errors:
		for index, error in enumerate(errors, start=1):
			print(f"\nFailure #{index}")
			print(f"File: {error['file']}")
			print(f"Error: {error['error']}")
			print("Traceback:")
			print(error["traceback"])
		print(f"Error log: {ERROR_LOG_FILE}")


if __name__ == "__main__":
	main()
