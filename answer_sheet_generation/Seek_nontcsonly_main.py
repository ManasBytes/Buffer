import os
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Protection, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.protection import SheetProtection

# ================== PATHS ==================
input_folder  = r"C:\Q_python\Seek_nontcs\Main\28 March 2026\Shift 1\METADATA"
output_folder = r"C:\Q_python\Seek_nontcs\Output"
log_file      = os.path.join(output_folder, "error_log.txt")

os.makedirs(output_folder, exist_ok=True)

# ================== HELPERS ==================

def get_engine(path):
    return 'xlrd' if path.lower().endswith('.xls') else 'openpyxl'


def find_config_sheet(path):
    xl = pd.ExcelFile(path, engine=get_engine(path))
    for s in xl.sheet_names:
        if "configuration" in s.lower():
            return s
    raise ValueError(f"No Configuration sheet found. Available: {xl.sheet_names}")


def clean_id(val):
    """Prevent scientific notation — return as plain string."""
    if pd.isna(val) or str(val).strip() in ('', 'nan', 'NaN'):
        return ''
    val = str(val).strip()
    if val.endswith('.0') and val[:-2].isdigit():
        val = val[:-2]
    return val


def clean_marks(val):
    if pd.isna(val) or str(val).strip() in ('', 'nan', 'NaN'):
        return ''
    val = str(val).strip()
    if val.endswith('.0') and val[:-2].isdigit():
        val = val[:-2]
    return val


def subject_to_filename(raw):
    """Convert subject string from col D to a safe output filename."""
    clean = re.sub(r'[:\-]', ' ', str(raw))       # remove colon and hyphen
    clean = re.sub(r'[\\/*?"<>|]', '', clean)      # remove other illegal chars
    clean = re.sub(r'\s+', ' ', clean).strip()     # collapse spaces
    return f"{clean} NON_TCS entry.xlsx"


# ================== PROCESS ==================
success_count = 0
errors = []

files = [
    f for f in os.listdir(input_folder)
    if f.lower().endswith(('.xlsx', '.xls'))
    and not f.startswith('~$')             # ← skip Windows temp lock files
]

print(f"Found {len(files)} file(s) to process.\n")

for filename in files:
    filepath = os.path.join(input_folder, filename)
    try:
        print(f" Processing: {filename}")

        # ── Detect sheet ─────────────────────────────────────────────
        sheet_name = find_config_sheet(filepath)
        print(f"   → Sheet: {sheet_name}")

        # ── Read data ────────────────────────────────────────────────
        df = pd.read_excel(filepath, sheet_name=sheet_name,
                           header=0, dtype=str, engine=get_engine(filepath))
        print(f"   → Rows: {len(df)}  Cols: {df.shape[1]}")

        if df.shape[1] < 5:
            raise ValueError(f"Need at least 5 columns, found {df.shape[1]}")

        # ── Output filename from Subject (col D = index 3, first data row) ──
        subject_raw = str(df.iloc[0, 3]).strip()
        if not subject_raw or subject_raw.lower() == 'nan':
            subject_raw = os.path.splitext(filename)[0]   # fallback to input filename
        output_name = subject_to_filename(subject_raw)
        output_path = os.path.join(output_folder, output_name)

        # ── Filter rows: skip COMPREHENSION and blank types ──────────
        rows_out = []
        sno = 0
        for _, row in df.iterrows():
            q_type = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ''
            if not q_type or q_type.upper() == 'COMPREHENSION':
                continue
            sno += 1
            rows_out.append({
                "Question ID's":    clean_id(row.iloc[0]),
                "S.No":             sno,
                "Question Types":   q_type,
                "Marks":            clean_marks(row.iloc[4]),
                "Enter Your Answer": ""
            })

        if not rows_out:
            raise ValueError("No non-Comprehension rows found after filtering.")

        final_df = pd.DataFrame(rows_out,
                                columns=["Question ID's", "S.No", "Question Types",
                                         "Marks", "Enter Your Answer"])

        # ── Build workbook ───────────────────────────────────────────
        wb = Workbook()
        ws = wb.active
        ws.title = "Question Paper Details"

        wrap = Alignment(wrap_text=True)

        for r_idx, row_data in enumerate(
                dataframe_to_rows(final_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.alignment = wrap
                # Force Question ID cells as text
                if c_idx == 1 and r_idx > 1:
                    cell.number_format = '@'

        # ── Column widths ────────────────────────────────────────────
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 25

        # ── Protection: A B C D locked | E editable ──────────────────
        for row_cells in ws.iter_rows():
            for cell in row_cells:
                if cell.column == 5:
                    cell.protection = Protection(locked=False)
                else:
                    cell.protection = Protection(locked=True)

        ws.protection = SheetProtection(sheet=True, password="1234")

        # ── Save ─────────────────────────────────────────────────────
        wb.save(output_path)
        print(f"    Saved: {output_name}  ({sno} rows)\n")
        success_count += 1

    except Exception as e:
        msg = f"{filename} → {e}"
        print(f"    Error: {msg}\n")
        errors.append(msg)

# ── Error log ────────────────────────────────────────────────────────
if errors:
    with open(log_file, 'w', encoding='utf-8') as f:
        f.write(f"Error Log — {len(errors)} failure(s)\n{'='*60}\n")
        for e in errors:
            f.write(e + "\n")
    print(f"  {len(errors)} error(s) logged to: {log_file}")

print(f"\n🏁 PROCESS COMPLETE")
print(f" Success: {success_count}")
print(f" Failed : {len(errors)}")
